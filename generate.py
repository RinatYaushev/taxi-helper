#!/usr/bin/env python3
"""
Генератор Excel-калькулятора прибутку в таксі (Uklon) з data.json.

Використання:
    python3 generate.py                # читає data.json → uklon_calculator.xlsx
    python3 generate.py my.json out.xlsx

Логіка комісії:
    Готівка / Комбінована : комісія Uklon = amount * commission_uklon_pct%
    Безготівка            : + додатково amount * commission_cashless_pct% (2% за вивід)
Газ (грн) = distance * (1 + empty_run_coef) * (gas_consumption/100 * gas_price)
Чистий    = amount - комісія - газ
"""

import json
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ---------- стилі ----------
HDR = PatternFill("solid", fgColor="2E5A88")
HDRF = Font(color="FFFFFF", bold=True)
CFG = PatternFill("solid", fgColor="FFF2CC")
GOOD = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")
SUBH = PatternFill("solid", fgColor="DDEBF7")
TITLE = Font(bold=True, size=14, color="2E5A88")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def compute(trip, s):
    """Порахувати газ, комісію, чистий, грн/км для однієї поїздки."""
    amount = float(trip["amount"])
    dist = float(trip["distance"])
    fuel_per_km = s["gas_consumption_l_100km"] / 100 * s["gas_price_per_l"]
    gas = dist * (1 + s["empty_run_coef"]) * fuel_per_km
    commission = amount * s["commission_uklon_pct"] / 100
    if trip["payment"] == "Безготівка":
        commission += amount * s["commission_cashless_pct"] / 100
    net = amount - commission - gas
    gross_per_km = amount / dist if dist else 0
    net_per_km = net / dist if dist else 0
    rating = "OK" if net_per_km >= s["threshold_net_per_km"] else "погана"

    marginal = s.get("marginal_net_per_km", 10)
    if net_per_km >= s["threshold_net_per_km"]:
        rec = "бери"
    elif net_per_km >= marginal:
        rec = "думай"
    else:
        rec = "пропускай"
    return {
        "gas": gas, "commission": commission, "net": net,
        "gross_per_km": gross_per_km, "net_per_km": net_per_km,
        "rating": rating, "rec": rec,
    }


def avg(values):
    vals = [v for v in values if v is not None]
    return sum(vals) / len(vals) if vals else 0


def build(data, out_path):
    s = data["settings"]
    trips = data["trips"]
    rows = [dict(t, **compute(t, s)) for t in trips]

    wb = Workbook()

    # ---------------- Налаштування ----------------
    ws = wb.active
    ws.title = "Налаштування"
    ws["A1"] = "Налаштування (джерело — data.json)"
    ws["A1"].font = TITLE
    cfg_rows = [
        ("Параметр", "Значення", "Одиниці / коментар"),
        ("Витрата ГБО (газ)", s["gas_consumption_l_100km"], "л/100 км"),
        ("Ціна газу", s["gas_price_per_l"], "грн/л"),
        ("Коеф. порожнього пробігу", s["empty_run_coef"], "частка порожняку від корисного км"),
        ("Комісія Uklon", s["commission_uklon_pct"], "%"),
        ("Комісія за вивід безготівки", s["commission_cashless_pct"], "%, лише для безготівки"),
        ("Поріг 'хорошої' поїздки", s["threshold_net_per_km"], "грн/км чистими"),
    ]
    for i, (a, b, c) in enumerate(cfg_rows, start=3):
        ws.cell(i, 1, a); ws.cell(i, 2, b); ws.cell(i, 3, c)
        if i == 3:
            for col in range(1, 4):
                ws.cell(i, col).fill = HDR; ws.cell(i, col).font = HDRF
        else:
            ws.cell(i, 2).fill = CFG; ws.cell(i, 2).font = BOLD
        for col in range(1, 4):
            ws.cell(i, col).border = BORDER
    fuel_per_km = s["gas_consumption_l_100km"] / 100 * s["gas_price_per_l"]
    ws["A11"] = "Собівартість палива, грн/км"
    ws["B11"] = round(fuel_per_km, 2)
    ws["A12"] = "Собівартість з порожняком, грн/км"
    ws["B12"] = round(fuel_per_km * (1 + s["empty_run_coef"]), 2)
    for r in (11, 12):
        ws.cell(r, 1).font = BOLD; ws.cell(r, 2).font = BOLD
    ws["A14"] = "⚠ Ці значення генеруються з data.json. Прав JSON і перезапусти generate.py."
    ws["A14"].font = Font(italic=True, color="B00000")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40

    # ---------------- Поїздки ----------------
    t = wb.create_sheet("Поїздки")
    headers = ["Дата/час", "Оплата", "Сума, грн", "Відстань, км", "Подача",
               "Призначення", "Зона", "Газ, грн", "Комісія, грн", "Чистий, грн",
               "грн/км", "Чистий грн/км", "Оцінка", "Рекоменд."]
    for j, h in enumerate(headers, start=1):
        cell = t.cell(1, j, h)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CENTER; cell.border = BORDER
    for i, r in enumerate(rows, start=2):
        t.cell(i, 1, r["datetime"])
        t.cell(i, 2, r["payment"])
        t.cell(i, 3, round(r["amount"], 2))
        t.cell(i, 4, round(r["distance"], 2))
        t.cell(i, 5, r.get("from", ""))
        t.cell(i, 6, r.get("to", ""))
        t.cell(i, 7, r.get("zone", ""))
        t.cell(i, 8, round(r["gas"], 2))
        t.cell(i, 9, round(r["commission"], 2))
        t.cell(i, 10, round(r["net"], 2))
        t.cell(i, 11, round(r["gross_per_km"], 2))
        t.cell(i, 12, round(r["net_per_km"], 2))
        t.cell(i, 13, r["rating"])
        t.cell(i, 14, r["rec"])
    last = len(rows) + 1
    # бордери + вирівнювання
    for i in range(2, last + 1):
        for j in range(1, 15):
            cell = t.cell(i, j)
            cell.border = BORDER
            if j in (2, 3, 4, 7, 8, 9, 10, 11, 12, 13, 14):
                cell.alignment = CENTER
            if j in (3, 4, 8, 9, 10, 11, 12):
                cell.number_format = "0.00"
    # підсвітка оцінки
    if last >= 2:
        t.conditional_formatting.add(f"M2:M{last}",
            CellIsRule(operator="equal", formula=['"OK"'], fill=GOOD))
        t.conditional_formatting.add(f"M2:M{last}",
            CellIsRule(operator="equal", formula=['"погана"'], fill=BAD))
        # підсвітка рекомендації
        t.conditional_formatting.add(f"N2:N{last}",
            CellIsRule(operator="equal", formula=['"бери"'], fill=GOOD))
        t.conditional_formatting.add(f"N2:N{last}",
            CellIsRule(operator="equal", formula=['"думай"'], fill=CFG))
        t.conditional_formatting.add(f"N2:N{last}",
            CellIsRule(operator="equal", formula=['"пропускай"'], fill=BAD))
    for j, w in enumerate([13, 12, 10, 12, 30, 30, 12, 10, 11, 11, 9, 13, 10, 11], start=1):
        t.column_dimensions[get_column_letter(j)].width = w
    t.freeze_panes = "A2"

    # ---------------- Аналіз ----------------
    a = wb.create_sheet("Аналіз")
    a["A1"] = "Аналіз поїздок"; a["A1"].font = TITLE

    def by(pred):
        return [r for r in rows if pred(r)]

    total_amount = sum(r["amount"] for r in rows)
    total_gas = sum(r["gas"] for r in rows)
    total_comm = sum(r["commission"] for r in rows)
    total_net = sum(r["net"] for r in rows)
    total_km = sum(r["distance"] for r in rows)

    cash = by(lambda r: r["payment"] == "Готівка")
    cashless = by(lambda r: r["payment"] == "Безготівка")
    combo = by(lambda r: r["payment"] == "Комбінована")
    city = by(lambda r: r.get("zone") == "Місто")
    dead = by(lambda r: r.get("zone") == "Глухий кут")
    bad = by(lambda r: r["rating"] == "погана")
    ok = by(lambda r: r["rating"] == "OK")

    block = [
        ("ЗАГАЛЬНЕ", None),
        ("К-сть поїздок", len(rows)),
        ("Сума валова, грн", round(total_amount, 2)),
        ("Комісія всього, грн", round(total_comm, 2)),
        ("Витрати на газ, грн", round(total_gas, 2)),
        ("ЧИСТИЙ прибуток, грн", round(total_net, 2)),
        ("Пробіг клієнтів, км", round(total_km, 2)),
        ("Середній ЧИСТИЙ грн/км", round(avg([r["net_per_km"] for r in rows]), 2)),
        ("Чистий грн/км (загальний)", round(total_net / total_km, 2) if total_km else 0),
        ("Середній чистий за поїздку, грн", round(avg([r["net"] for r in rows]), 2)),
        ("Частка газу у виручці, %", round(total_gas / total_amount * 100, 1) if total_amount else 0),
        ("Частка комісії у виручці, %", round(total_comm / total_amount * 100, 1) if total_amount else 0),
        ("", None),
        ("ЗА ОПЛАТОЮ (сер. чистий грн/км)", None),
        ("Готівка", round(avg([r["net_per_km"] for r in cash]), 2)),
        ("Безготівка", round(avg([r["net_per_km"] for r in cashless]), 2)),
        ("Комбінована", round(avg([r["net_per_km"] for r in combo]), 2)),
        ("", None),
        ("ЗА ЗОНОЮ", None),
        ("Місто — к-сть", len(city)),
        ("Місто — сер. чистий грн/км", round(avg([r["net_per_km"] for r in city]), 2)),
        ("Глухий кут — к-сть", len(dead)),
        ("Глухий кут — сер. чистий грн/км", round(avg([r["net_per_km"] for r in dead]), 2)),
        ("", None),
        ("ЯКІСТЬ ПОТОКУ", None),
        ("OK поїздок", len(ok)),
        ("Поганих (нижче порогу)", len(bad)),
        ("Частка поганих, %", round(len(bad) / len(rows) * 100, 1) if rows else 0),
    ]
    r = 3
    for label, value in block:
        a.cell(r, 1, label)
        if value is None:
            a.cell(r, 1).font = BOLD
            a.cell(r, 1).fill = SUBH
        else:
            a.cell(r, 2, value); a.cell(r, 2).font = BOLD
        r += 1
    a.column_dimensions["A"].width = 34
    a.column_dimensions["B"].width = 14

    # ---------------- Фільтри (пороги для Автопілота) ----------------
    f = wb.create_sheet("Фільтри")
    f["A1"] = "Фільтри для Автопілота Uklon"
    f["A1"].font = TITLE

    flt = s.get("filters", {})
    fuel_km = s["gas_consumption_l_100km"] / 100 * s["gas_price_per_l"]
    breakeven = fuel_km * (1 + s["empty_run_coef"])  # собівартість з порожняком
    norm_mult = flt.get("normal_price_km_mult", 1.6)
    long_mult = flt.get("long_price_km_mult", 1.4)
    norm_min_pk = round(breakeven * norm_mult, 1)
    long_min_pk = round(breakeven * long_mult, 1)

    f["A3"] = "Собівартість палива, грн/км"
    f["B3"] = round(fuel_km, 2)
    f["A4"] = "Собівартість з порожняком (беззбитковість), грн/км"
    f["B4"] = round(breakeven, 2)
    for r in (3, 4):
        f.cell(r, 1).font = BOLD
        f.cell(r, 2).font = BOLD
        f.cell(r, 2).fill = CFG

    # ЗОЛОТЕ ПРАВИЛО: мінімальна валова ціна за км, щоб вийти на поріг
    thr = s["threshold_net_per_km"]
    cash_rate = s["commission_uklon_pct"] / 100
    min_gross = (thr + breakeven) / (1 - cash_rate)
    f["A5"] = "★ ЗОЛОТЕ ПРАВИЛО: мін. валова ціна, грн/км"
    f["B5"] = round(min_gross, 1)
    f["A5"].font = Font(bold=True, size=12, color="B00000")
    f["B5"].font = Font(bold=True, size=12, color="B00000")
    f["B5"].fill = GOOD
    f["C5"] = "= сума ÷ км. Нижче — не бери!"
    f["C5"].font = Font(italic=True)

    # таблиця трьох профілів
    f["A6"] = "Три профілі рішення (приймати/пропускати)"
    f["A6"].font = BOLD
    f["A6"].fill = SUBH
    prof_hdr = ["Тип", "Довжина, км", "Макс. подача, км", "Мін. грн/км", "Кінцева точка"]
    for j, hh in enumerate(prof_hdr, start=1):
        c = f.cell(7, j, hh)
        c.fill = HDR
        c.font = HDRF
        c.alignment = CENTER
        c.border = BORDER
    prof_rows = [
        ("Коротка", f"до {flt.get('short_max_km', 3)}",
         f"≤ {flt.get('short_max_pickup_km', 1.5)}", "висока (базовий тариф)", "будь-яка в місті"),
        ("Звичайна", f"{flt.get('short_max_km', 3)}–{flt.get('normal_max_km', 10)}",
         f"≤ {flt.get('normal_max_pickup_km', 3)}", f"≥ {norm_min_pk}", "НЕ глухий кут"),
        ("Довга", f"{flt.get('normal_max_km', 10)}+",
         f"≤ {flt.get('long_max_pickup_km', 5)}", f"≥ {long_min_pk}", "НЕ село / тупик"),
    ]
    for i, pr in enumerate(prof_rows, start=8):
        for j, val in enumerate(pr, start=1):
            c = f.cell(i, j, val)
            c.border = BORDER
            if j != 1 and j != 5:
                c.alignment = CENTER

    # готові значення Автопілота
    f["A12"] = "Що виставити в Автопілоті"
    f["A12"].font = BOLD
    f["A12"].fill = SUBH
    ap = [
        ("Мінімальна вартість замовлення, грн", flt.get("autopilot_min_order", 90)),
        ("Макс. відстань подачі, км", flt.get("normal_max_pickup_km", 3)),
        ("Готівка + Безготівка", "приймати обидві"),
        ("Фільтр 'Мені по дорозі'", "тримати на кінець зміни (уникати тупиків)"),
    ]
    for i, (a_, b_) in enumerate(ap, start=13):
        f.cell(i, 1, a_)
        f.cell(i, 2, b_)
        f.cell(i, 2).font = BOLD
        f.cell(i, 2).fill = CFG

    # глухі кути / живі зони
    dead = ", ".join(s.get("dead_end_areas", []))
    live = ", ".join(s.get("live_areas", []))
    f["A18"] = "Глухі кути (обережно з довгими туди):"
    f["A18"].font = BOLD
    f["B18"] = dead
    f["A19"] = "Живі зони (гарне повернення):"
    f["A19"].font = BOLD
    f["B19"] = live
    f["A21"] = ("Правило: довгу поїздку (10+ км) береш тільки якщо кінцева — "
                "жива зона або грн/км явно вище порогу.")
    f["A21"].font = Font(italic=True, color="B00000")

    f.column_dimensions["A"].width = 42
    for col in "BCDE":
        f.column_dimensions[col].width = 20

    # ---------------- Інструкція ----------------
    h = wb.create_sheet("Інструкція")
    h["A1"] = "Як користуватись"; h["A1"].font = TITLE
    lines = [
        "1. Усі дані — у файлі data.json (налаштування + список поїздок).",
        "2. Додаєш нову поїздку — вписуєш об'єкт у масив 'trips':",
        '   {"datetime":"17.08 19:30","payment":"Готівка","amount":150,',
        '    "distance":5.2,"from":"...","to":"...","zone":"Місто"}',
        "   payment: Готівка | Безготівка | Комбінована",
        "   zone:    Місто | Глухий кут",
        "3. Запускаєш: python3 generate.py",
        "4. Відкриваєш свіжий uklon_calculator.xlsx.",
        "",
        "Щоб змінити ставки/поріг — прав секцію 'settings' у data.json.",
        "Глухий кут = напрямок, звідки важко взяти зворотне замовлення.",
    ]
    for i, ln in enumerate(lines, start=3):
        h.cell(i, 1, ln)
    h.column_dimensions["A"].width = 80

    wb.save(out_path)
    return len(rows), total_net


def main():
    in_path = sys.argv[1] if len(sys.argv) > 1 else "data.json"
    out_path = sys.argv[2] if len(sys.argv) > 2 else f"{date.today():%Y-%m-%d}-result.xlsx"
    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)
    n, net = build(data, out_path)
    print(f"OK: {n} поїздок → {out_path} | чистий прибуток: {net:.2f} грн")


if __name__ == "__main__":
    main()
